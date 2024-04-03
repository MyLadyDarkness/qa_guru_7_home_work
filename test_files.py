from zipfile import ZipFile

import pypdf
from openpyxl import load_workbook

from conftest import ARCH


def test_pdf():
    with ZipFile(ARCH) as zip_pdf:
        with zip_pdf.open("Python Notes.pdf") as ext_pdf:
            reader = pypdf.PdfReader(ext_pdf)
            pdf_content = reader.pages[1].extract_text()
    assert "If - flow control" in pdf_content


def test_xlsx():
    with ZipFile(ARCH) as zip_xlsx:
        with zip_xlsx.open("example.xlsx") as example_xlsx:
            reader = load_workbook(example_xlsx)
            xls_file = reader.active
    assert "Some" in xls_file.cell(row=1, column=1).value


def test_csv():
    with ZipFile(ARCH) as zip_csv:
        with zip_csv.open("example3.csv") as example3_csv:
            csv_content = example3_csv.read().decode('utf-8')
    assert "work3" in csv_content
